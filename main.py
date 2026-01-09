"""
CSV to XLSX Converter
A beautiful Windows 11 utility for converting CSV files to Excel format
Author: Your Name
License: MIT
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os
import threading
from pathlib import Path


class CSVtoXLSXConverter(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        # Window configuration
        self.title("CSV to XLSX Converter")
        self.geometry("900x700")
        self.minsize(800, 600)
        
        # Set appearance
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        
        # Variables
        self.files_list = []
        self.output_folder = ctk.StringVar(value=str(Path.home() / "Documents"))
        self.encoding_var = ctk.StringVar(value="utf-8")
        self.delimiter_var = ctk.StringVar(value=",")
        self.header_row_var = ctk.BooleanVar(value=True)
        self.auto_width_var = ctk.BooleanVar(value=True)
        self.freeze_header_var = ctk.BooleanVar(value=True)
        self.add_filters_var = ctk.BooleanVar(value=True)
        self.style_header_var = ctk.BooleanVar(value=True)
        self.zebra_stripes_var = ctk.BooleanVar(value=False)
        self.header_color_var = ctk.StringVar(value="#4A90D9")
        self.text_color_var = ctk.StringVar(value="#FFFFFF")
        self.skip_rows_var = ctk.StringVar(value="0")
        self.sheet_name_var = ctk.StringVar(value="Sheet1")
        
        self.create_widgets()
        
    def create_widgets(self):
        # Main container with padding
        self.main_frame = ctk.CTkFrame(self, corner_radius=0)
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header
        self.create_header()
        
        # Create tabview for organized settings
        self.tabview = ctk.CTkTabview(self.main_frame, height=450)
        self.tabview.pack(fill="both", expand=True, pady=(10, 10))
        
        self.tabview.add("📁 Files")
        self.tabview.add("⚙️ CSV Settings")
        self.tabview.add("🎨 Excel Styling")
        self.tabview.add("📊 Preview")
        
        self.create_files_tab()
        self.create_csv_settings_tab()
        self.create_styling_tab()
        self.create_preview_tab()
        
        # Bottom action bar
        self.create_action_bar()
        
    def create_header(self):
        header_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        header_frame.pack(fill="x", pady=(0, 10))
        
        title_label = ctk.CTkLabel(
            header_frame, 
            text="📊 CSV to XLSX Converter",
            font=ctk.CTkFont(size=28, weight="bold")
        )
        title_label.pack(side="left")
        
        subtitle_label = ctk.CTkLabel(
            header_frame,
            text="Convert your CSV files to beautiful Excel spreadsheets",
            font=ctk.CTkFont(size=14),
            text_color="gray"
        )
        subtitle_label.pack(side="left", padx=(20, 0))
        
        # Theme toggle
        self.theme_switch = ctk.CTkSwitch(
            header_frame,
            text="Dark Mode",
            command=self.toggle_theme,
            onvalue="dark",
            offvalue="light"
        )
        self.theme_switch.select()
        self.theme_switch.pack(side="right")
        
    def create_files_tab(self):
        files_tab = self.tabview.tab("📁 Files")
        
        # File selection frame
        file_frame = ctk.CTkFrame(files_tab)
        file_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Buttons row
        btn_frame = ctk.CTkFrame(file_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(10, 10))
        
        self.add_files_btn = ctk.CTkButton(
            btn_frame,
            text="➕ Add Files",
            command=self.add_files,
            width=140
        )
        self.add_files_btn.pack(side="left", padx=5)
        
        self.add_folder_btn = ctk.CTkButton(
            btn_frame,
            text="📁 Add Folder",
            command=self.add_folder,
            width=140
        )
        self.add_folder_btn.pack(side="left", padx=5)
        
        self.remove_btn = ctk.CTkButton(
            btn_frame,
            text="🗑️ Remove Selected",
            command=self.remove_selected,
            width=140,
            fg_color="#D94444",
            hover_color="#B33636"
        )
        self.remove_btn.pack(side="left", padx=5)
        
        self.clear_btn = ctk.CTkButton(
            btn_frame,
            text="🧹 Clear All",
            command=self.clear_files,
            width=140,
            fg_color="#666666",
            hover_color="#555555"
        )
        self.clear_btn.pack(side="left", padx=5)
        
        # Files listbox with scrollbar
        list_frame = ctk.CTkFrame(file_frame)
        list_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.files_listbox = ctk.CTkTextbox(list_frame, height=200)
        self.files_listbox.pack(fill="both", expand=True)
        
        # File count label
        self.file_count_label = ctk.CTkLabel(
            file_frame,
            text="No files selected",
            font=ctk.CTkFont(size=12)
        )
        self.file_count_label.pack(pady=5)
        
        # Output folder selection
        output_frame = ctk.CTkFrame(file_frame, fg_color="transparent")
        output_frame.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(
            output_frame,
            text="📂 Output Folder:",
            font=ctk.CTkFont(size=14, weight="bold")
        ).pack(side="left")
        
        self.output_entry = ctk.CTkEntry(
            output_frame,
            textvariable=self.output_folder,
            width=400
        )
        self.output_entry.pack(side="left", padx=10)
        
        ctk.CTkButton(
            output_frame,
            text="Browse",
            command=self.browse_output,
            width=100
        ).pack(side="left")
        
    def create_csv_settings_tab(self):
        csv_tab = self.tabview.tab("⚙️ CSV Settings")
        
        settings_frame = ctk.CTkFrame(csv_tab)
        settings_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Encoding
        encoding_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        encoding_frame.pack(fill="x", pady=10, padx=20)
        
        ctk.CTkLabel(
            encoding_frame,
            text="📝 Encoding:",
            font=ctk.CTkFont(size=14, weight="bold"),
            width=150,
            anchor="w"
        ).pack(side="left")
        
        encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1251", "cp1252", "iso-8859-1", "ascii"]
        self.encoding_menu = ctk.CTkOptionMenu(
            encoding_frame,
            variable=self.encoding_var,
            values=encodings,
            width=200
        )
        self.encoding_menu.pack(side="left", padx=10)
        
        # Delimiter
        delimiter_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        delimiter_frame.pack(fill="x", pady=10, padx=20)
        
        ctk.CTkLabel(
            delimiter_frame,
            text="🔤 Delimiter:",
            font=ctk.CTkFont(size=14, weight="bold"),
            width=150,
            anchor="w"
        ).pack(side="left")
        
        delimiters = [(",", "Comma (,)"), (";", "Semicolon (;)"), ("\t", "Tab"), ("|", "Pipe (|)")]
        self.delimiter_menu = ctk.CTkOptionMenu(
            delimiter_frame,
            variable=self.delimiter_var,
            values=[d[0] for d in delimiters],
            width=200
        )
        self.delimiter_menu.pack(side="left", padx=10)
        
        # Skip rows
        skip_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        skip_frame.pack(fill="x", pady=10, padx=20)
        
        ctk.CTkLabel(
            skip_frame,
            text="⏭️ Skip Rows:",
            font=ctk.CTkFont(size=14, weight="bold"),
            width=150,
            anchor="w"
        ).pack(side="left")
        
        self.skip_rows_entry = ctk.CTkEntry(
            skip_frame,
            textvariable=self.skip_rows_var,
            width=200
        )
        self.skip_rows_entry.pack(side="left", padx=10)
        
        # Sheet name
        sheet_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        sheet_frame.pack(fill="x", pady=10, padx=20)
        
        ctk.CTkLabel(
            sheet_frame,
            text="📋 Sheet Name:",
            font=ctk.CTkFont(size=14, weight="bold"),
            width=150,
            anchor="w"
        ).pack(side="left")
        
        self.sheet_name_entry = ctk.CTkEntry(
            sheet_frame,
            textvariable=self.sheet_name_var,
            width=200
        )
        self.sheet_name_entry.pack(side="left", padx=10)
        
        # Checkboxes
        checkbox_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        checkbox_frame.pack(fill="x", pady=20, padx=20)
        
        self.header_checkbox = ctk.CTkCheckBox(
            checkbox_frame,
            text="First row contains headers",
            variable=self.header_row_var
        )
        self.header_checkbox.pack(anchor="w", pady=5)
        
    def create_styling_tab(self):
        style_tab = self.tabview.tab("🎨 Excel Styling")
        
        style_frame = ctk.CTkFrame(style_tab)
        style_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Left column - checkboxes
        left_frame = ctk.CTkFrame(style_frame, fg_color="transparent")
        left_frame.pack(side="left", fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(
            left_frame,
            text="📐 Layout Options",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", pady=(0, 15))
        
        self.auto_width_checkbox = ctk.CTkCheckBox(
            left_frame,
            text="Auto-fit column width",
            variable=self.auto_width_var
        )
        self.auto_width_checkbox.pack(anchor="w", pady=5)
        
        self.freeze_checkbox = ctk.CTkCheckBox(
            left_frame,
            text="Freeze header row",
            variable=self.freeze_header_var
        )
        self.freeze_checkbox.pack(anchor="w", pady=5)
        
        self.filter_checkbox = ctk.CTkCheckBox(
            left_frame,
            text="Add auto-filters",
            variable=self.add_filters_var
        )
        self.filter_checkbox.pack(anchor="w", pady=5)
        
        self.style_header_checkbox = ctk.CTkCheckBox(
            left_frame,
            text="Style header row",
            variable=self.style_header_var
        )
        self.style_header_checkbox.pack(anchor="w", pady=5)
        
        self.zebra_checkbox = ctk.CTkCheckBox(
            left_frame,
            text="Zebra stripes (alternating row colors)",
            variable=self.zebra_stripes_var
        )
        self.zebra_checkbox.pack(anchor="w", pady=5)
        
        # Right column - colors
        right_frame = ctk.CTkFrame(style_frame, fg_color="transparent")
        right_frame.pack(side="right", fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(
            right_frame,
            text="🎨 Color Settings",
            font=ctk.CTkFont(size=16, weight="bold")
        ).pack(anchor="w", pady=(0, 15))
        
        # Header color
        header_color_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        header_color_frame.pack(fill="x", pady=10)
        
        ctk.CTkLabel(
            header_color_frame,
            text="Header Background:",
            width=150,
            anchor="w"
        ).pack(side="left")
        
        self.header_color_entry = ctk.CTkEntry(
            header_color_frame,
            textvariable=self.header_color_var,
            width=100
        )
        self.header_color_entry.pack(side="left", padx=10)
        
        self.header_color_preview = ctk.CTkButton(
            header_color_frame,
            text="",
            width=30,
            height=30,
            fg_color=self.header_color_var.get(),
            hover_color=self.header_color_var.get(),
            command=lambda: self.pick_color("header")
        )
        self.header_color_preview.pack(side="left")
        
        # Text color
        text_color_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        text_color_frame.pack(fill="x", pady=10)
        
        ctk.CTkLabel(
            text_color_frame,
            text="Header Text Color:",
            width=150,
            anchor="w"
        ).pack(side="left")
        
        self.text_color_entry = ctk.CTkEntry(
            text_color_frame,
            textvariable=self.text_color_var,
            width=100
        )
        self.text_color_entry.pack(side="left", padx=10)
        
        self.text_color_preview = ctk.CTkButton(
            text_color_frame,
            text="",
            width=30,
            height=30,
            fg_color=self.text_color_var.get(),
            hover_color=self.text_color_var.get(),
            command=lambda: self.pick_color("text")
        )
        self.text_color_preview.pack(side="left")
        
        # Preset themes
        preset_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        preset_frame.pack(fill="x", pady=20)
        
        ctk.CTkLabel(
            preset_frame,
            text="Quick Presets:",
            font=ctk.CTkFont(weight="bold")
        ).pack(anchor="w", pady=(0, 10))
        
        presets_btn_frame = ctk.CTkFrame(preset_frame, fg_color="transparent")
        presets_btn_frame.pack(fill="x")
        
        presets = [
            ("Blue", "#4A90D9", "#FFFFFF"),
            ("Green", "#4CAF50", "#FFFFFF"),
            ("Orange", "#FF9800", "#000000"),
            ("Purple", "#9C27B0", "#FFFFFF"),
            ("Dark", "#333333", "#FFFFFF"),
        ]
        
        for name, bg, fg in presets:
            btn = ctk.CTkButton(
                presets_btn_frame,
                text=name,
                width=60,
                height=28,
                fg_color=bg,
                text_color=fg,
                command=lambda b=bg, f=fg: self.apply_preset(b, f)
            )
            btn.pack(side="left", padx=3)
            
    def create_preview_tab(self):
        preview_tab = self.tabview.tab("📊 Preview")
        
        preview_frame = ctk.CTkFrame(preview_tab)
        preview_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Preview controls
        controls_frame = ctk.CTkFrame(preview_frame, fg_color="transparent")
        controls_frame.pack(fill="x", pady=10, padx=10)
        
        self.preview_btn = ctk.CTkButton(
            controls_frame,
            text="🔄 Load Preview",
            command=self.load_preview
        )
        self.preview_btn.pack(side="left")
        
        ctk.CTkLabel(
            controls_frame,
            text="(Shows first 100 rows of selected file)",
            text_color="gray"
        ).pack(side="left", padx=10)
        
        # Preview text area
        self.preview_text = ctk.CTkTextbox(preview_frame, height=300)
        self.preview_text.pack(fill="both", expand=True, padx=10, pady=10)
        
    def create_action_bar(self):
        action_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        action_frame.pack(fill="x", pady=(10, 0))
        
        # Progress bar
        self.progress = ctk.CTkProgressBar(action_frame, height=20)
        self.progress.pack(fill="x", pady=(0, 10))
        self.progress.set(0)
        
        # Status label
        self.status_label = ctk.CTkLabel(
            action_frame,
            text="Ready to convert",
            font=ctk.CTkFont(size=12)
        )
        self.status_label.pack(pady=(0, 10))
        
        # Convert button
        self.convert_btn = ctk.CTkButton(
            action_frame,
            text="🚀 Convert to XLSX",
            command=self.start_conversion,
            height=50,
            font=ctk.CTkFont(size=18, weight="bold"),
            fg_color="#28a745",
            hover_color="#218838"
        )
        self.convert_btn.pack(fill="x")
        
    # ========== Functionality Methods ==========
    
    def toggle_theme(self):
        if self.theme_switch.get() == "dark":
            ctk.set_appearance_mode("dark")
        else:
            ctk.set_appearance_mode("light")
            
    def add_files(self):
        files = filedialog.askopenfilenames(
            title="Select CSV Files",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        for f in files:
            if f not in self.files_list:
                self.files_list.append(f)
        self.update_files_display()
        
    def add_folder(self):
        folder = filedialog.askdirectory(title="Select Folder with CSV Files")
        if folder:
            for file in Path(folder).glob("*.csv"):
                if str(file) not in self.files_list:
                    self.files_list.append(str(file))
        self.update_files_display()
        
    def remove_selected(self):
        # For simplicity, we'll clear and let user re-add
        # In a full implementation, you'd track selection
        messagebox.showinfo("Info", "Use 'Clear All' to remove files, then re-add the ones you need.")
        
    def clear_files(self):
        self.files_list = []
        self.update_files_display()
        
    def update_files_display(self):
        self.files_listbox.delete("1.0", "end")
        for i, f in enumerate(self.files_list, 1):
            self.files_listbox.insert("end", f"{i}. {f}\n")
        
        count = len(self.files_list)
        self.file_count_label.configure(
            text=f"{count} file(s) selected" if count > 0 else "No files selected"
        )
        
    def browse_output(self):
        folder = filedialog.askdirectory(title="Select Output Folder")
        if folder:
            self.output_folder.set(folder)
            
    def pick_color(self, color_type):
        # Simple color input - in production, use a color picker dialog
        pass
        
    def apply_preset(self, bg, fg):
        self.header_color_var.set(bg)
        self.text_color_var.set(fg)
        self.header_color_preview.configure(fg_color=bg, hover_color=bg)
        self.text_color_preview.configure(fg_color=fg, hover_color=fg)
        
    def load_preview(self):
        if not self.files_list:
            messagebox.showwarning("Warning", "Please add files first!")
            return
            
        try:
            file_path = self.files_list[0]
            skip_rows = int(self.skip_rows_var.get()) if self.skip_rows_var.get() else 0
            
            df = pd.read_csv(
                file_path,
                encoding=self.encoding_var.get(),
                delimiter=self.delimiter_var.get(),
                skiprows=skip_rows,
                nrows=100
            )
            
            self.preview_text.delete("1.0", "end")
            self.preview_text.insert("1.0", df.to_string())
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not load preview: {str(e)}")
            
    def start_conversion(self):
        if not self.files_list:
            messagebox.showwarning("Warning", "Please add files to convert!")
            return
            
        self.convert_btn.configure(state="disabled")
        thread = threading.Thread(target=self.convert_files)
        thread.start()
        
    def convert_files(self):
        total = len(self.files_list)
        success = 0
        errors = []
        
        for i, file_path in enumerate(self.files_list):
            try:
                self.update_status(f"Converting {os.path.basename(file_path)}...")
                self.progress.set((i + 1) / total)
                
                # Read CSV
                skip_rows = int(self.skip_rows_var.get()) if self.skip_rows_var.get() else 0
                df = pd.read_csv(
                    file_path,
                    encoding=self.encoding_var.get(),
                    delimiter=self.delimiter_var.get(),
                    skiprows=skip_rows
                )
                
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = self.sheet_name_var.get() or "Sheet1"
                
                # Write data
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=self.header_row_var.get()), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply styling
                self.apply_excel_styling(ws, df)
                
                # Save
                output_name = Path(file_path).stem + ".xlsx"
                output_path = Path(self.output_folder.get()) / output_name
                wb.save(output_path)
                
                success += 1
                
            except Exception as e:
                errors.append(f"{os.path.basename(file_path)}: {str(e)}")
                
        self.update_status(f"Completed! {success}/{total} files converted.")
        self.progress.set(1)
        self.convert_btn.configure(state="normal")
        
        if errors:
            messagebox.showwarning("Conversion Completed with Errors", "\n".join(errors))
        else:
            messagebox.showinfo("Success", f"Successfully converted {success} file(s)!")
            
    def apply_excel_styling(self, ws, df):
        # Header styling
        if self.style_header_var.get() and self.header_row_var.get():
            header_fill = PatternFill(
                start_color=self.header_color_var.get().replace("#", ""),
                end_color=self.header_color_var.get().replace("#", ""),
                fill_type="solid"
            )
            header_font = Font(
                bold=True,
                color=self.text_color_var.get().replace("#", "")
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        # Auto-fit columns
        if self.auto_width_var.get():
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        # Freeze header
        if self.freeze_header_var.get():
            ws.freeze_panes = "A2"
            
        # Add filters
        if self.add_filters_var.get():
            ws.auto_filter.ref = ws.dimensions
            
        # Zebra stripes
        if self.zebra_stripes_var.get():
            stripe_fill = PatternFill(
                start_color="F5F5F5",
                end_color="F5F5F5",
                fill_type="solid"
            )
            start_row = 2 if self.header_row_var.get() else 1
            for row_idx in range(start_row, ws.max_row + 1):
                if row_idx % 2 == 0:
                    for cell in ws[row_idx]:
                        cell.fill = stripe_fill
                        
        # Add borders
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                
    def update_status(self, message):
        self.status_label.configure(text=message)
        self.update()


def main():
    app = CSVtoXLSXConverter()
    app.mainloop()


if __name__ == "__main__":
    main()
